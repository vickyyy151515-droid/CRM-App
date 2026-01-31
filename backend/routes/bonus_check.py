# Customer Bonus Check Routes

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
import uuid
import io
import csv
from .deps import User, get_db, get_current_user, get_admin_user, get_jakarta_now
from datetime import datetime, timedelta

router = APIRouter(tags=["Bonus Check"])


class BonusCheckSubmission(BaseModel):
    customer_id: str
    product_id: str


class BonusCheckResponse(BaseModel):
    success: bool
    message: str
    submission_id: Optional[str] = None


@router.post("/bonus-check/submit", response_model=BonusCheckResponse)
async def submit_bonus_check(data: BonusCheckSubmission, user: User = Depends(get_current_user)):
    """
    Staff submits a customer for bonus eligibility check.
    System validates:
    1. Customer exists in staff's reserved member list
    2. Reserved member is not expired (30-day rule)
    """
    db = get_db()
    
    if user.role != 'staff':
        raise HTTPException(status_code=403, detail="Only staff can submit bonus checks")
    
    now = get_jakarta_now()
    customer_id_normalized = data.customer_id.strip().upper()
    
    # Get product info
    product = await db.products.find_one({'id': data.product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check if customer is in staff's reserved member list
    # Check both customer_id and customer_name fields for backwards compatibility
    reserved = await db.reserved_members.find_one({
        '$or': [
            {'customer_id': {'$regex': f'^{customer_id_normalized}$', '$options': 'i'}},
            {'customer_name': {'$regex': f'^{customer_id_normalized}$', '$options': 'i'}}
        ],
        'reserved_by': user.id,
        'product_id': data.product_id,
        'status': 'approved'
    }, {'_id': 0})
    
    if not reserved:
        return BonusCheckResponse(
            success=False,
            message=f"Customer '{data.customer_id}' tidak ditemukan di daftar Reserved Member Anda untuk produk {product['name']}. Pastikan customer sudah di-reserve dan sudah diapprove."
        )
    
    # Check expiration based on 30-day grace period
    # Get grace period config
    config = await db.reserved_member_config.find_one({'type': 'cleanup_config'}, {'_id': 0})
    global_grace_days = 30  # Default
    
    if config:
        # Check for product-specific override
        product_overrides = config.get('product_overrides', {})
        if data.product_id in product_overrides:
            global_grace_days = product_overrides[data.product_id]
        else:
            global_grace_days = config.get('global_grace_days', 30)
    
    # Calculate expiration
    approved_at = reserved.get('approved_at') or reserved.get('created_at')
    if approved_at:
        if isinstance(approved_at, str):
            try:
                approved_date = datetime.fromisoformat(approved_at.replace('Z', '+00:00'))
            except ValueError:
                approved_date = now
        else:
            approved_date = approved_at
        
        expiration_date = approved_date + timedelta(days=global_grace_days)
        
        # Make now timezone-aware if needed for comparison
        now_aware = now
        if expiration_date.tzinfo is not None and now_aware.tzinfo is None:
            from datetime import timezone
            now_aware = now.replace(tzinfo=timezone.utc)
        elif expiration_date.tzinfo is None and now_aware.tzinfo is not None:
            expiration_date = expiration_date.replace(tzinfo=now_aware.tzinfo)
        
        if now_aware > expiration_date:
            return BonusCheckResponse(
                success=False,
                message=f"Reserved member untuk customer '{data.customer_id}' sudah expired (lebih dari {global_grace_days} hari). Silakan ajukan reserved member baru."
            )
    
    # Check for duplicate submission this month
    current_month = now.strftime('%Y-%m')
    existing = await db.bonus_check_submissions.find_one({
        'customer_id_normalized': customer_id_normalized,
        'product_id': data.product_id,
        'staff_id': user.id,
        'month': current_month
    })
    
    if existing:
        return BonusCheckResponse(
            success=False,
            message=f"Customer '{data.customer_id}' sudah disubmit untuk cek bonus bulan ini."
        )
    
    # Create submission
    submission_id = str(uuid.uuid4())
    submission = {
        'id': submission_id,
        'customer_id': data.customer_id.strip(),
        'customer_id_normalized': customer_id_normalized,
        'product_id': data.product_id,
        'product_name': product['name'],
        'staff_id': user.id,
        'staff_name': user.name,
        'reserved_member_id': reserved.get('id'),
        'month': current_month,
        'submitted_at': now.isoformat(),
        'status': 'pending'  # Admin can review later
    }
    
    await db.bonus_check_submissions.insert_one(submission)
    
    return BonusCheckResponse(
        success=True,
        message=f"Customer '{data.customer_id}' berhasil disubmit untuk cek bonus bulan {current_month}.",
        submission_id=submission_id
    )


@router.get("/bonus-check/my-submissions")
async def get_my_bonus_submissions(
    month: Optional[str] = None,
    user: User = Depends(get_current_user)
):
    """Get staff's own bonus check submissions"""
    db = get_db()
    
    if user.role != 'staff':
        raise HTTPException(status_code=403, detail="Only staff can access this endpoint")
    
    query = {'staff_id': user.id}
    
    if month:
        query['month'] = month
    else:
        # Default to current month
        now = get_jakarta_now()
        query['month'] = now.strftime('%Y-%m')
    
    submissions = await db.bonus_check_submissions.find(
        query, {'_id': 0}
    ).sort('submitted_at', -1).to_list(1000)
    
    return {
        'submissions': submissions,
        'total': len(submissions),
        'month': query['month']
    }


@router.get("/bonus-check/products")
async def get_products_for_bonus_check(user: User = Depends(get_current_user)):
    """Get list of products for bonus check dropdown"""
    db = get_db()
    products = await db.products.find({}, {'_id': 0}).to_list(100)
    return products


@router.get("/bonus-check/admin/all")
async def get_all_bonus_submissions(
    month: Optional[str] = None,
    staff_id: Optional[str] = None,
    product_id: Optional[str] = None,
    user: User = Depends(get_admin_user)
):
    """Admin: Get all bonus check submissions"""
    db = get_db()
    
    query = {}
    
    if month:
        query['month'] = month
    else:
        # Default to current month
        now = get_jakarta_now()
        query['month'] = now.strftime('%Y-%m')
    
    if staff_id:
        query['staff_id'] = staff_id
    
    if product_id:
        query['product_id'] = product_id
    
    submissions = await db.bonus_check_submissions.find(
        query, {'_id': 0}
    ).sort('submitted_at', -1).to_list(10000)
    
    # Group by staff for summary
    staff_summary = {}
    for sub in submissions:
        sid = sub['staff_id']
        if sid not in staff_summary:
            staff_summary[sid] = {
                'staff_id': sid,
                'staff_name': sub['staff_name'],
                'count': 0
            }
        staff_summary[sid]['count'] += 1
    
    return {
        'submissions': submissions,
        'total': len(submissions),
        'month': query['month'],
        'by_staff': list(staff_summary.values())
    }


@router.get("/bonus-check/admin/export")
async def export_bonus_submissions(
    month: Optional[str] = None,
    staff_id: Optional[str] = None,
    product_id: Optional[str] = None,
    format: str = "csv",
    user: User = Depends(get_admin_user)
):
    """Admin: Export bonus check submissions to CSV or Excel"""
    db = get_db()
    
    query = {}
    
    if month:
        query['month'] = month
    else:
        now = get_jakarta_now()
        query['month'] = now.strftime('%Y-%m')
    
    if staff_id:
        query['staff_id'] = staff_id
    
    if product_id:
        query['product_id'] = product_id
    
    submissions = await db.bonus_check_submissions.find(
        query, {'_id': 0}
    ).sort('submitted_at', -1).to_list(10000)
    
    if format == "csv":
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['No', 'Customer ID', 'Product', 'Staff Name', 'Submitted At', 'Month', 'Status'])
        
        # Data rows
        for idx, sub in enumerate(submissions, 1):
            writer.writerow([
                idx,
                sub.get('customer_id', ''),
                sub.get('product_name', ''),
                sub.get('staff_name', ''),
                sub.get('submitted_at', ''),
                sub.get('month', ''),
                sub.get('status', 'pending')
            ])
        
        output.seek(0)
        
        filename = f"bonus_check_{query['month']}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:
        # Excel format using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Bonus Check {query['month']}"
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            
            # Headers
            headers = ['No', 'Customer ID', 'Product', 'Staff Name', 'Submitted At', 'Month', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for idx, sub in enumerate(submissions, 1):
                ws.cell(row=idx+1, column=1, value=idx)
                ws.cell(row=idx+1, column=2, value=sub.get('customer_id', ''))
                ws.cell(row=idx+1, column=3, value=sub.get('product_name', ''))
                ws.cell(row=idx+1, column=4, value=sub.get('staff_name', ''))
                ws.cell(row=idx+1, column=5, value=sub.get('submitted_at', ''))
                ws.cell(row=idx+1, column=6, value=sub.get('month', ''))
                ws.cell(row=idx+1, column=7, value=sub.get('status', 'pending'))
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"bonus_check_{query['month']}.xlsx"
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            raise HTTPException(status_code=400, detail="Excel export not available. Please use CSV format.")


@router.get("/bonus-check/admin/staff-list")
async def get_staff_list_for_filter(user: User = Depends(get_admin_user)):
    """Get list of staff for filter dropdown"""
    db = get_db()
    staff = await db.users.find(
        {'role': 'staff'},
        {'_id': 0, 'id': 1, 'name': 1, 'email': 1}
    ).to_list(1000)
    return staff
